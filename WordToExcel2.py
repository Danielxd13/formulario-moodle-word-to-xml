from docx import Document
from openpyxl import Workbook
from datetime import datetime
import argparse


def letra_a_numero(letra):
    """Convierte la letra de la respuesta a un número"""
    conversion = {'A': 1, 'B': 2, 'C': 3, 'D': 4}
    # Extraer la primera letra de la respuesta (ignorando el paréntesis)
    primera_letra = letra.strip()[0] if letra else ''
    return conversion.get(primera_letra, '')

def limpiar_texto(texto, tipo='pregunta'):
    """Limpia el texto removiendo números iniciales, guiones y letras con paréntesis"""
    if not texto:
        return texto

    if tipo == 'pregunta':
        # Busca el primer punto y espacio después del número
        if '.' in texto and texto[0].isdigit():
            inicio = texto.find('.') + 1
            texto_limpio = texto[inicio:].strip()
            # Eliminar guiones al inicio si existen
            if texto_limpio.startswith('-'):
                texto_limpio = texto_limpio.lstrip('- ')
            return texto_limpio
    else:  # tipo == 'respuesta'
        # Remueve A), B), C), D) del inicio
        if texto.startswith(('A)', 'B)', 'C)', 'D)')):
            texto_limpio = texto[2:].strip()
            
            # Para debug: agregar esto después de la línea texto_limpio = texto[2:].strip()
            print(f"Caracteres en texto: {[ord(c) for c in texto_limpio]}")
            
            # Lista de posibles caracteres de viñetas de Word
            caracteres_vineta = ['•', '·', '○', '●', '■', '□', '◦', '▪', '▫', '-', '*']
            
            # Detectar si hay viñetas
            es_lista = any(texto_limpio.find(vineta) != -1 for vineta in caracteres_vineta)
            
            if es_lista:
                # Dividir por cualquiera de los caracteres de viñeta
                partes = texto_limpio
                for vineta in caracteres_vineta:
                    if vineta in partes:
                        partes = partes.replace(vineta, '|SPLIT|')
                
                # Dividir y limpiar las partes
                items = [p.strip() for p in partes.split('|SPLIT|')]
                # Filtrar elementos vacíos y limpiar espacios
                return [item for item in items if item and not item.isspace()]
            
            # Eliminar guiones al inicio si existen
            if texto_limpio.startswith('-'):
                texto_limpio = texto_limpio.lstrip('- ')
            return texto_limpio
    return texto.strip()

def leer_docx(ruta_archivo):
    try:
        doc = Document(ruta_archivo)
        preguntas = []
        respuestas = []
        respuestas_correctas = []
        pregunta_actual = ""
        respuestas_actuales = ['', '', '', '']
        respuesta_correcta_actual = ''

        for parrafo in doc.paragraphs:
            # Verificar si el párrafo tiene viñetas en su formato XML
            tiene_vinetas = False
            if hasattr(parrafo._element, 'pPr') and parrafo._element.pPr is not None:
                if parrafo._element.pPr.numPr is not None:
                    tiene_vinetas = True

            texto = parrafo.text.strip()
            
            if texto:
                # Si empieza con número y punto, es una pregunta
                if texto[0].isdigit() and ('. ' in texto or '.' in texto):
                    # Guardar la pregunta, respuestas y respuesta correcta anteriores si existen
                    if pregunta_actual:
                        preguntas.append(limpiar_texto(pregunta_actual, 'pregunta'))
                        respuestas.append(respuestas_actuales)
                        respuestas_correctas.append(respuesta_correcta_actual)
                    # Iniciar nueva pregunta
                    pregunta_actual = texto
                    respuestas_actuales = ['', '', '', '']
                    respuesta_correcta_actual = ''
                # Si empieza con A), B), C) o D), es una respuesta
                elif any(texto.startswith(prefix) for prefix in ['A)', 'B)', 'C)', 'D)']):
                    # Procesar cada línea de respuesta
                    for run in parrafo.runs:
                        # Verificar si el texto está resaltado
                        if run.font.highlight_color:
                            # Limpiar y convertir la respuesta correcta a número
                            texto_resaltado = run.text.strip()
                            if texto_resaltado:
                                respuesta_correcta_actual = letra_a_numero(texto_resaltado)
                            break
                    
                    # Si el párrafo tiene viñetas, dividir el texto
                    if tiene_vinetas:
                        # Encontrar qué letra de respuesta es (A, B, C o D)
                        letra_respuesta = texto[0]
                        indice_respuesta = ord(letra_respuesta) - ord('A')
                        
                        # Dividir el texto en elementos de lista
                        elementos = [item.strip() for item in texto[2:].split('\n') if item.strip()]
                        respuestas_actuales[indice_respuesta] = elementos
                    else:
                        # Procesar como antes para respuestas sin viñetas
                        lineas = texto.split('\n')
                        for linea in lineas:
                            linea = linea.strip()
                            if linea:
                                if linea.startswith('A)'):
                                    respuestas_actuales[0] = limpiar_texto(linea, 'respuesta')
                                elif linea.startswith('B)'):
                                    respuestas_actuales[1] = limpiar_texto(linea, 'respuesta')
                                elif linea.startswith('C)'):
                                    respuestas_actuales[2] = limpiar_texto(linea, 'respuesta')
                                elif linea.startswith('D)'):
                                    respuestas_actuales[3] = limpiar_texto(linea, 'respuesta')

        # Guardar la última pregunta y sus respuestas
        if pregunta_actual:
            preguntas.append(limpiar_texto(pregunta_actual, 'pregunta'))
            respuestas.append(respuestas_actuales[:])
            respuestas_correctas.append(respuesta_correcta_actual)
        
        return preguntas, respuestas, respuestas_correctas

    except Exception as e:
        print(f"Error al leer el archivo: {str(e)}")
        return None, None, None

def crear_excel(preguntas, respuestas, respuestas_correctas=None, nombre_archivo=None):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Contenido DOCX"
        
        # Agregar encabezado
        ws['A1'] = "ID"
        ws['B1'] = "Preguntas"
        ws['C1'] = "Respuesta1"
        ws['D1'] = "Respuesta2"
        ws['E1'] = "Respuesta3"
        ws['F1'] = "Respuesta4"
        ws['G1'] = "Respuesta correcta"

        # Agregar contenido
        fila_actual = 2
        for idx, (pregunta, resp) in enumerate(zip(preguntas, respuestas), start=1):
            ws[f'A{fila_actual}'] = idx
            ws[f'B{fila_actual}'] = pregunta
            
            tiene_subitems = False
            max_subitems = 1
            
            # Procesar respuestas y contar máximo de subitems
            respuestas_procesadas = []
            for respuesta in resp:
                if isinstance(respuesta, list):
                    tiene_subitems = True
                    max_subitems = max(max_subitems, len(respuesta))
                    respuestas_procesadas.append(respuesta)
                else:
                    respuestas_procesadas.append([respuesta] if respuesta else [''])

            # Si hay subitems, crear filas adicionales
            for subitem in range(max_subitems):
                for j, respuesta in enumerate(respuestas_procesadas):
                    if subitem < len(respuesta):
                        ws[f'{chr(67+j)}{fila_actual}'] = respuesta[subitem]
                    else:
                        ws[f'{chr(67+j)}{fila_actual}'] = ''
                
                # Agregar la respuesta correcta solo en la primera fila del grupo
                if subitem == 0 and respuestas_correctas and idx-1 < len(respuestas_correctas):
                    ws[f'G{fila_actual}'] = respuestas_correctas[idx-1]
                
                if tiene_subitems and subitem < max_subitems - 1:
                    fila_actual += 1
                    # Copiar el ID y la pregunta en las filas adicionales
                    ws[f'A{fila_actual}'] = idx
                    ws[f'B{fila_actual}'] = pregunta
            
            fila_actual += 1
        
        # Si no se proporciona nombre, crear uno con la fecha actual
        if nombre_archivo is None:
            fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
            nombre_archivo = f"contenido_docx_{fecha}.xlsx"
        
        # Guardar el archivo
        wb.save(nombre_archivo)
        return nombre_archivo
    
    except Exception as e:
        print(f"Error al crear el archivo Excel: {e}")
        return None

if __name__ == "__main__":
    # Crear el parser de argumentos
    parser = argparse.ArgumentParser(description='Procesa un archivo DOCX y genera un Excel con preguntas y respuestas.')
    parser.add_argument('ruta_archivo', nargs='?', 
                       help='Ruta al archivo DOCX que se va a procesar',
                       default="SIMULACRO 1. CELADORES.docx")
    parser.add_argument('--salida', '-s', 
                       help='Nombre del archivo Excel de salida (opcional)',
                       default="Contenido_docx.xlsx")
    
    # Parsear los argumentos
    args = parser.parse_args()
    
    # Usar los argumentos
    preguntas, respuestas, respuestas_correctas = leer_docx(args.ruta_archivo)
    
    if preguntas and respuestas:
        excel_generado = crear_excel(preguntas, respuestas, respuestas_correctas, args.salida)
        if excel_generado:
            print(f"\nArchivo Excel creado exitosamente: {excel_generado}")
