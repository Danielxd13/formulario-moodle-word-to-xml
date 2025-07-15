from docx import Document
from openpyxl import Workbook
from datetime import datetime
import argparse


def letra_a_numero(letra):
    """Convierte la letra de la respuesta a un número"""
    conversion = {'A': 1, 'B': 2, 'C': 3, 'D': 4}
    # Extraer la primera letra de la respuesta (ignorando el paréntesis)
    primera_letra = letra.strip()[0] if letra else ''
    return conversion.get(primera_letra, '')

def limpiar_texto(texto, tipo='pregunta'):
    """Limpia el texto removiendo números iniciales, guiones y letras con paréntesis"""
    if tipo == 'pregunta':
        # Buscar el primer dígito
        if texto[0].isdigit():
            # Encontrar el final del formato de numeración
            for i, char in enumerate(texto):
                if char in ['.', ')', '-'] and i > 0:
                    # Si encuentra un guión o punto después del número
                    inicio = i + 1
                    # Si hay un espacio después del separador, saltar el espacio
                    if inicio < len(texto) and texto[inicio] == ' ':
                        inicio += 1
                    texto_limpio = texto[inicio:].strip()
                    # Eliminar guiones adicionales al inicio si existen
                    if texto_limpio.startswith('-'):
                        texto_limpio = texto_limpio.lstrip('- ')
                    return texto_limpio
    else:  # tipo == 'respuesta'
        # Remueve A), B), C), D) del inicio
        if texto.startswith(('A)', 'B)', 'C)', 'D)', 'a)', 'b)', 'c)', 'd)', 'o A) ', 'o B) ', 'o C) ', 'o D) ', 'o a) ', 'o b) ', 'o c) ', 'o d) ')):
            texto_limpio = texto[2:].strip()
            # Eliminar guiones al inicio si existen
            if texto_limpio.startswith('-'):
                texto_limpio = texto_limpio.lstrip('- ')
            return texto_limpio
    return texto.strip()

def leer_docx(ruta_archivo):
    try:
        doc = Document(ruta_archivo)
        preguntas = []
        respuestas = []
        respuestas_correctas = []  # Nueva lista para almacenar las respuestas correctas
        pregunta_actual = ""
        respuestas_actuales = ['', '', '', '']
        respuesta_correcta_actual = ''  # Variable para almacenar la respuesta correcta
        respuesta_idx = 0  # Inicializar el índice de respuesta fuera del bucle

        for parrafo in doc.paragraphs:
            texto = parrafo.text.strip()
            
            if texto:
                # Detectar si es una viñeta o numeración automática
                es_vineta = parrafo.style.name.lower().startswith('list') or (parrafo._p.pPr is not None and parrafo._p.pPr.numPr is not None)

                # Verificar si es una pregunta con diferentes formatos de numeración o lista numerada
                es_pregunta = False
                if texto[0].isdigit():
                    for separador in ['. ', '.', ')', '.-']:
                        if separador in texto[:4]:
                            es_pregunta = True
                            break
                # Si es una lista numerada y no una viñeta de respuestas, también es pregunta
                if es_vineta and not any(texto.startswith(prefix) for prefix in ['A)', 'B)', 'C)', 'D)', 'a)', 'b)', 'c)', 'd)', 'o A)', 'o B)', 'o C)', 'o D)', 'o a)', 'o b)', 'o c)', 'o d)']):
                    es_pregunta = True

                if es_pregunta:
                    if pregunta_actual:
                        preguntas.append(limpiar_texto(pregunta_actual, 'pregunta'))
                        respuestas.append(respuestas_actuales)
                        respuestas_correctas.append(respuesta_correcta_actual)
                    pregunta_actual = texto
                    respuestas_actuales = ['', '', '', '']
                    respuesta_correcta_actual = ''
                    respuesta_idx = 0  # Reiniciar el índice para las respuestas
                elif any(texto.startswith(prefix) for prefix in ['A)', 'B)', 'C)', 'D)', 'a)', 'b)', 'c)', 'd)', 'o A)', 'o B)', 'o C)', 'o D)', 'o a)', 'o b)', 'o c)', 'o d)']) or es_vineta:
                    # Si es viñeta o respuesta clásica, agregar como respuesta
                    if respuesta_idx < 4:
                        respuestas_actuales[respuesta_idx] = limpiar_texto(texto, 'respuesta')
                        # Detección de respuesta correcta igual que antes...
                        for run in parrafo.runs:
                            if (run.font.highlight_color or run.font.bold or run.font.underline):
                                respuesta_correcta_actual = respuesta_idx + 1  # +1 para que sea 1,2,3,4
                        respuesta_idx += 1

        # Guardar la última pregunta y sus respuestas
        if pregunta_actual:
            preguntas.append(limpiar_texto(pregunta_actual, 'pregunta'))
            respuestas.append(respuestas_actuales[:])
            respuestas_correctas.append(respuesta_correcta_actual)
        
        return preguntas, respuestas, respuestas_correctas

    except Exception as e:
        print(f"Error al leer el archivo: {str(e)}")
        return None, None, None

def crear_excel(preguntas, respuestas, respuestas_correctas=None, nombre_archivo=None):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Contenido DOCX"
        
        # Agregar encabezado
        ws['A1'] = "ID"
        ws['B1'] = "Preguntas"
        ws['C1'] = "Respuesta1"
        ws['D1'] = "Respuesta2"
        ws['E1'] = "Respuesta3"
        ws['F1'] = "Respuesta4"
        ws['G1'] = "Respuesta correcta"

        # Agregar contenido
        for i, (pregunta, resp) in enumerate(zip(preguntas, respuestas), start=2):
            ws[f'A{i}'] = i - 1
            ws[f'B{i}'] = pregunta
            for j, respuesta in enumerate(resp):
                ws[f'{chr(67+j)}{i}'] = respuesta
            
            # Agregar la respuesta correcta si existe
            if respuestas_correctas and i-2 < len(respuestas_correctas):
                ws[f'G{i}'] = respuestas_correctas[i-2]
        
        # Si no se proporciona nombre, crear uno con la fecha actual
        if nombre_archivo is None:
            fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
            nombre_archivo = f"contenido_docx_{fecha}.xlsx"
        
        # Guardar el archivo
        wb.save(nombre_archivo)
        return nombre_archivo
    
    except Exception as e:
        print(f"Error al crear el archivo Excel: {e}")
        return None

if __name__ == "__main__":
    # Crear el parser de argumentos
    parser = argparse.ArgumentParser(description='Procesa un archivo DOCX y genera un Excel con preguntas y respuestas.')
    parser.add_argument('ruta_archivo', help='Ruta al archivo DOCX que se va a procesar')
    parser.add_argument('--salida', '-s', help='Nombre del archivo Excel de salida (opcional)', default="Contenido_docx.xlsx")
    
    # Parsear los argumentos
    args = parser.parse_args()
    
    # Usar los argumentos
    preguntas, respuestas, respuestas_correctas = leer_docx(args.ruta_archivo)
    
    if preguntas and respuestas:
        excel_generado = crear_excel(preguntas, respuestas, respuestas_correctas, args.salida)
        if excel_generado:
            print(f"\nArchivo Excel creado exitosamente: {excel_generado}")
